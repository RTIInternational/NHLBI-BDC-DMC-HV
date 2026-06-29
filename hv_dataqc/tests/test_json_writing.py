"""Tests for the JSON-safe writer in hv_dataqc_common."""

from __future__ import annotations

import json

from hv_dataqc.hv_dataqc_common import (
    json_safe,
    write_json_atomic,
    write_xlsx,
    XLSX_FMT_COUNT,
    XLSX_FMT_DECIMAL,
)


def test_json_safe_nulls_non_finite_floats():
    assert json_safe({"inf": float("inf"), "nan": float("nan"), "ok": 2.5}) == {
        "inf": None, "nan": None, "ok": 2.5,
    }


def test_json_safe_stringifies_non_str_keys():
    assert json_safe({1: "a", 2.0: "b", "s": "c"}) == {"1": "a", "2.0": "b", "s": "c"}


def test_json_safe_sets_and_tuples_to_lists():
    out = json_safe({"set": {3, 1, 2}, "tuple": (4, 5)})
    assert sorted(out["set"]) == [1, 2, 3]
    assert out["tuple"] == [4, 5]


def test_write_json_atomic_roundtrips(tmp_path):
    data = {"k": [1, 2, {"nan": float("nan")}], 9: "intkey"}
    out = tmp_path / "out.json"
    write_json_atomic(out, data)
    loaded = json.loads(out.read_text())
    assert loaded == {"k": [1, 2, {"nan": None}], "9": "intkey"}
    assert not (tmp_path / "out.json.tmp").exists()


def test_write_xlsx_writes_numbers_with_formats(tmp_path):
    from openpyxl import load_workbook

    out = tmp_path / "t.xlsx"
    write_xlsx(
        out,
        headers=["variable", "n", "mean"],
        rows=[["AST SGOT", "15584", "27.5"], ["BMI", "", ""]],
        column_formats=[None, XLSX_FMT_COUNT, XLSX_FMT_DECIMAL],
        sheet_title="Table S5",
    )
    ws = load_workbook(out).active
    assert ws.title == "Table S5"
    assert ws.freeze_panes == "A2"
    # Numeric strings are coerced to real numbers and carry the column format.
    assert ws["B2"].value == 15584 and ws["B2"].number_format == XLSX_FMT_COUNT
    assert ws["C2"].value == 27.5 and ws["C2"].number_format == XLSX_FMT_DECIMAL
    # The text column is left as text.
    assert ws["A2"].value == "AST SGOT"
    # Empty cells stay empty (None), not 0.
    assert ws["B3"].value is None
