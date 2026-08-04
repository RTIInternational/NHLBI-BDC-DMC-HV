"""Shared helpers for HV-DataQC scripts.

All helpers operate on aggregate metadata only; no participant-level rows are
written or logged by this module.
"""

from __future__ import annotations

import ast
import json
import math
import re
from pathlib import Path
from typing import Any, Callable
from xml.etree import ElementTree as ET

import pandas as pd


LogFn = Callable[[str], None]


def canonical_phv_id(raw_id: str) -> str:
    """Return canonical PHV accession: lower-case, version suffix stripped."""
    return str(raw_id or "").split(".")[0].lower()


def json_safe(value: Any) -> Any:
    """Recursively convert values to strict-JSON-safe structures."""
    if isinstance(value, dict):
        return {str(k): json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [json_safe(v) for v in value]
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    return value


def write_json_atomic(
    path: Path,
    data: Any,
    *,
    ensure_ascii: bool = False,
    default: Callable[[Any], Any] | None = str,
) -> None:
    """Write strict JSON via temp file then atomic replace.

    Sanitizes the structure (``json_safe``) then streams it to disk.  Creates
    parent directories as needed.

    Note: very large structures hold a sanitized copy briefly during the
    ``json_safe`` pass.  Callers producing huge artifacts (e.g. the source
    extractor) should bound the structure itself — for example by not emitting
    the multi-PHV joint-distribution crosstabs when they are not needed —
    rather than relying on the writer to stream a multi-gigabyte dict.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f"{path.name}.tmp")
    try:
        with tmp_path.open("w", encoding="utf-8") as fh:
            json.dump(
                json_safe(data),
                fh,
                indent=2,
                ensure_ascii=ensure_ascii,
                default=default,
                allow_nan=False,
            )
        tmp_path.replace(path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


# Number-format strings (Excel) shared by the S4/S5 xlsx writers.
XLSX_FMT_COUNT = "#,##0"        # integer counts with thousands separators
XLSX_FMT_DECIMAL = "#,##0.00"  # 2-decimal stats with thousands separators


def coerce_number(value: Any) -> Any:
    """Return value as int/float if it is numeric (or a numeric string), else as-is.

    Writing real numbers (not pre-formatted strings) is what lets the cell
    number-format apply, so the reader gets commas / fixed decimals without any
    manual Excel formatting.
    """
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip()
    if s == "":
        return None
    try:
        f = float(s.replace(",", ""))
    except ValueError:
        return value
    return int(f) if f.is_integer() else f


def write_xlsx(
    path: Path,
    headers: list[str],
    rows: "list[list[Any]]",
    *,
    column_formats: "list[str | None] | None" = None,
    sheet_title: str = "Sheet1",
    freeze_header: bool = True,
) -> None:
    """Write a single-sheet .xlsx with per-column number formats.

    Numeric cells are written as numbers and given *column_formats[i]* (an Excel
    format string, or None to leave as text/general), so thousands separators
    and fixed decimals come from the cell format — no manual Excel step. Columns
    are auto-sized to their content. Creates parent dirs; writes atomically.

    Args:
        headers: column header labels (row 1).
        rows: list of row value lists, aligned to *headers*.
        column_formats: optional per-column Excel number-format string; None for
            a column means leave values as-is (text). Shorter lists are padded
            with None.
        sheet_title: worksheet name.
        freeze_header: freeze row 1 so the header stays visible when scrolling.
    """
    from openpyxl import Workbook  # local import: optional dependency

    fmts = list(column_formats or [])
    fmts += [None] * (len(headers) - len(fmts))

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)

    widths = [len(str(h)) for h in headers]
    for row in rows:
        out_cells: list[Any] = []
        for ci, raw in enumerate(row):
            val = coerce_number(raw) if fmts[ci] is not None else raw
            out_cells.append(val)
            shown = "" if val is None else str(val)
            if len(shown) > widths[ci]:
                widths[ci] = len(shown)
        ws.append(out_cells)
        for ci, fmt in enumerate(fmts):
            if fmt is not None:
                cell = ws.cell(row=ws.max_row, column=ci + 1)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = fmt

    from openpyxl.utils import get_column_letter
    for ci, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(ci + 1)].width = min(max(w + 2, 8), 60)
    if freeze_header:
        ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f"{path.name}.tmp")
    try:
        wb.save(tmp_path)
        tmp_path.replace(path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def load_phv_name_map(
    cache_dir: Path,
    *,
    info: LogFn | None = None,
    warning: LogFn | None = None,
) -> dict[str, str]:
    """Load PHV-accession -> variable-name map from dbGaP data-dict XML files."""
    phv_names: dict[str, str] = {}
    pheno_dir = cache_dir / "pheno_variable_summaries"
    if not pheno_dir.exists():
        if info:
            info(f"cache pheno_variable_summaries/ not found at {pheno_dir} -- PHV names unavailable")
        return phv_names

    files = list(pheno_dir.glob("*.data_dict.xml"))
    if info:
        info(f"Loading PHV names from {len(files)} data_dict.xml files...")
    for dd_file in files:
        try:
            tree = ET.parse(dd_file)
            for var in tree.getroot().findall(".//variable"):
                phv_id = canonical_phv_id(var.get("id", ""))
                name = (var.findtext("name") or "").strip()
                if phv_id and name:
                    phv_names[phv_id] = name
        except ET.ParseError as exc:
            if warning:
                warning(f"Could not parse PHV name XML {dd_file.name}: {exc}")

    if info:
        info(f"PHV name map: {len(phv_names)} entries")
    return phv_names


def categorical_stats(series: pd.Series) -> dict[str, Any]:
    """Return value distribution for a categorical series."""
    n_total = int(len(series))
    n_missing = int(series.isna().sum())
    n_valid = n_total - n_missing
    counts = series.value_counts(dropna=True, sort=True)
    distribution: dict[str, dict[str, Any]] = {}
    for val, cnt in counts.items():
        distribution[str(val)] = {
            "n": int(cnt),
            "pct": round(100.0 * int(cnt) / n_valid, 2) if n_valid > 0 else 0.0,
        }

    return {
        "type": "categorical",
        "n_total": n_total,
        "n_valid": n_valid,
        "n_missing": n_missing,
        "pct_missing": round(n_missing / n_total * 100, 2) if n_total > 0 else 0.0,
        "n_distinct": int(series.nunique(dropna=True)),
        "distribution": distribution,
    }


def continuous_stats(series: pd.Series) -> dict[str, Any]:
    """Return descriptive statistics for a continuous (numeric) series."""
    numeric = pd.to_numeric(series, errors="coerce")
    n_total = int(len(numeric))
    s = numeric.dropna()
    n_valid = int(len(s))
    n_missing = n_total - n_valid
    result: dict[str, Any] = {
        "type": "continuous",
        "n_total": n_total,
        "n_valid": n_valid,
        "n_missing": n_missing,
        "pct_missing": round(n_missing / n_total * 100, 2) if n_total > 0 else 0.0,
        "n_distinct": int(s.nunique()),
    }
    if n_valid > 0:
        p25 = round(float(s.quantile(0.25)), 4)
        p75 = round(float(s.quantile(0.75)), 4)
        result.update(
            {
                "mean": round(float(s.mean()), 4),
                "sd": round(float(s.std()), 4),
                "median": round(float(s.median()), 4),
                "p5": round(float(s.quantile(0.05)), 4),
                "p25": p25,
                "p75": p75,
                "p95": round(float(s.quantile(0.95)), 4),
                "min": round(float(s.min()), 4),
                "max": round(float(s.max()), 4),
                "q1": p25,
                "q3": p75,
            }
        )
    else:
        result.update(
            {
                k: None
                for k in ["mean", "sd", "median", "p5", "p25", "p75", "p95", "min", "max", "q1", "q3"]
            }
        )
    return result


_INTEGER_FLOAT_RE = re.compile(r"^-?\d+\.0$")


def normalize_category_key(value: Any) -> str:
    """Normalize category keys emitted as numeric strings, JSON arrays, or Python reprs."""
    text = str(value).strip()

    parsed: Any | None = None
    if (text.startswith("[") and text.endswith("]")) or (text.startswith("(") and text.endswith(")")):
        for parser in (json.loads, ast.literal_eval):
            try:
                parsed = parser(text)
                break
            except (ValueError, SyntaxError, json.JSONDecodeError):
                parsed = None
        if isinstance(parsed, (list, tuple)) and len(parsed) == 1:
            text = str(parsed[0]).strip()

    if _INTEGER_FLOAT_RE.match(text):
        return text[:-2]
    return text
