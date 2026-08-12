#!/usr/bin/env python
"""Compare a generated Table S4 against the published supplementary sheet.

The published S4 is a *historical artifact*, not an oracle. It was produced by a
superseded pipeline that read curator spreadsheets and filtered phvs against
hand-made allow lists; both of those filters were retired on 2026-08-12 as
unmaintainable, so **exact agreement is not the goal** and counting rules must
never be tuned to hit its numbers. See ``README.md``.

Expect the generated table to read *higher* — the retired filters only ever
removed phvs. An increase needs no explanation. What this is for: spotting the
changes that go the other way. A cell that is blank here but populated there, or
that *drops* by an order of magnitude, is a generator bug until proven
otherwise.

Usage:
    ./.venv/bin/python transform_assessment/compare_s4_to_published.py \\
        --generated ~/Downloads/latest_s4_report.xlsx \\
        --published "~/Downloads/Data Harmonization Supplementary Data (6).xlsx"

    # only rows whose name matches a substring, with per-cohort detail
    ... --filter "alcohol" --detail
"""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

SHEET = "Table S4"
FIRST_DATA_ROW = 5
# Columns 2..21 are 10 (phv, n) pairs: 9 cohorts + TOTALS.
LAST_DATA_COL = 21
COHORTS = ["ARIC", "CARDIA", "CHS", "COPDGene", "FHS",
           "HCHS/SOL", "JHS", "MESA", "WHI", "TOTALS"]


def load_rows(path: Path) -> dict[str, list[int | None]]:
    """Label -> 20 cell values (phv, n per cohort). Later rows win ties."""
    ws = openpyxl.load_workbook(path, data_only=True)[SHEET]
    rows: dict[str, list[int | None]] = {}
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not isinstance(label, str) or not label.strip():
            continue
        label = label.strip()
        # The trailing note row is prose, not a variable.
        if len(label) > 80:
            continue
        vals = [ws.cell(r, c).value for c in range(2, LAST_DATA_COL + 1)]
        rows[label] = [int(v) if isinstance(v, (int, float)) else None for v in vals]
    return rows


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--generated", required=True, type=Path)
    ap.add_argument("--published", required=True, type=Path)
    ap.add_argument("--filter", default="", help="only rows containing this substring")
    ap.add_argument("--detail", action="store_true", help="per-cohort cell detail")
    args = ap.parse_args()

    gen = load_rows(args.generated.expanduser())
    pub = load_rows(args.published.expanduser())

    common = [k for k in gen if k in pub]
    if args.filter:
        needle = args.filter.lower()
        common = [k for k in common if needle in k.lower()]

    identical = lower = higher = equal_cells = 0
    blank_in_gen = pub_populated = 0
    offenders: list[tuple[int, str, str]] = []

    for label in common:
        p, g = pub[label], gen[label]
        if p == g:
            identical += 1
        # phv is every even index (0, 2, ... 18); n is odd.
        for i in range(0, 18, 2):
            pv, gv = p[i], g[i]
            if pv is not None:
                pub_populated += 1
                if gv is None:
                    blank_in_gen += 1
            if pv is None or gv is None:
                continue
            if gv > pv:
                higher += 1
            elif gv < pv:
                lower += 1
                if pv >= 10 and gv * 3 <= pv:
                    offenders.append((pv - gv, label, COHORTS[i // 2]))
            else:
                equal_cells += 1

    print(f"labels in both sheets : {len(common)}")
    print(f"rows identical        : {identical}")
    print(f"phv cells equal       : {equal_cells}")
    print(f"phv cells gen < pub   : {lower}")
    print(f"phv cells gen > pub   : {higher}")
    pct = 100 * blank_in_gen // max(pub_populated, 1)
    print(f"populated in pub, blank in gen : {blank_in_gen} of {pub_populated} ({pct}%)")

    if offenders:
        print("\nlargest shortfalls (pub >= 10 and gen <= pub/3):")
        for delta, label, cohort in sorted(offenders, reverse=True)[:20]:
            print(f"  -{delta:6}  {label[:44]:44} {cohort}")

    if args.detail:
        for label in common:
            p, g = pub[label], gen[label]
            if p == g:
                continue
            print(f"\n{label}")
            for i in range(0, 20, 2):
                if (p[i], p[i + 1]) == (g[i], g[i + 1]):
                    continue
                print(f"  {COHORTS[i // 2]:9} pub={str(p[i]):>6}/{str(p[i+1]):>9}"
                      f"   gen={str(g[i]):>6}/{str(g[i+1]):>9}")


if __name__ == "__main__":
    main()
