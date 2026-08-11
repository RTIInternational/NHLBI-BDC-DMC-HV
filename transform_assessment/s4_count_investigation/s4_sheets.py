"""Shared readers for Table S4 spreadsheets and pipeline CSVs.

Both scripts in this directory need the same two things: the published sheet
parsed into {label: {cohort: (phv, n)}}, and the same shape from the old
pipeline's CSV output.  Two normalizations are load-bearing, and skipping either
fabricates differences that are not in the data -- which is exactly what
happened before they were centralized here:

- A visually-empty cell is stored as '' in some exports and None in others.
  Comparing raw values reported 582 changed cells for 2026-03-23 -> 2026-06-25
  where there are in fact 4.
- The trailing TOTALS / TOTAL VARIABLES rows are summaries, not variables.
  Counting them made a version look like it had lost two variables.

Sheet layout: labels in col 1, cohort headers in row 3, phv/n subheaders in row
4, data from row 5, cohort column pairs from col 2.  The 2025-08-05 export has no
TOTALS column; later ones carry TOTALS at cols 20/21.
"""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

SHEET = "Table S4"
SUMMARY_ROWS = {"TOTALS", "TOTAL VARIABLES"}


def norm(v):
    """Normalize a cell: blank -> None, numeric -> float, else stripped str.

    Exports disagree on whether a blank is '' or None, and on whether a count is
    stored as int or float.  Neither difference is visible in the sheet, so
    neither may register as a change.
    """
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return s
    if isinstance(v, (int, float)):
        return float(v)
    return v


def load_xlsx(path: Path, drop_totals: bool = False):
    """Return (labels in sheet order, {label: {cohort: (phv, n)}}, cohorts).

    Labels come back as a list as well as a dict so callers can check row *order*
    and duplication, not just membership.  ``drop_totals`` excludes the TOTALS
    cohort column, which is a rollup rather than a cohort.
    """
    ws = openpyxl.load_workbook(path, data_only=True)[SHEET]
    cohorts = {ws.cell(row=3, column=c).value: c
               for c in range(2, ws.max_column + 1)
               if ws.cell(row=3, column=c).value}
    if drop_totals:
        cohorts.pop("TOTALS", None)
    labels: list[str] = []
    rows: dict[str, dict[str, tuple]] = {}
    for r in range(5, ws.max_row + 1):
        raw = ws.cell(row=r, column=1).value
        if raw is None or not str(raw).strip():
            continue
        label = str(raw).strip()
        if label in SUMMARY_ROWS:
            continue
        labels.append(label)
        rows.setdefault(label, {co: (norm(ws.cell(row=r, column=c0).value),
                                     norm(ws.cell(row=r, column=c0 + 1).value))
                                for co, c0 in cohorts.items()})
    return labels, rows, list(cohorts)


def load_csv(path: Path):
    """Read a pipeline CSV, whose columns are '<cohort>_phv' / '<cohort>_n'.

    Returns (labels in file order, {label: {cohort: (phv, n)}}) -- the same shape
    as ``load_xlsx`` so the two can be compared directly.
    """
    order: list[str] = []
    acc: dict[str, dict[str, dict[str, object]]] = {}
    with open(path, newline="") as fh:
        for rec in csv.DictReader(fh):
            label = (rec["variable"] or "").strip()
            if not label or label in SUMMARY_ROWS:
                continue
            order.append(label)
            acc.setdefault(label, {})
            for key, val in rec.items():
                if key == "variable" or key is None:
                    continue
                cohort, _, part = key.rpartition("_")
                acc[label].setdefault(cohort, {})[part] = norm(val)
    rows = {label: {co: (d.get("phv"), d.get("n")) for co, d in cols.items()}
            for label, cols in acc.items()}
    return order, rows


def fmt(cell: tuple) -> str:
    """Render a (phv, n) pair for display; '-' for an empty cell.

    Counts are whole numbers stored as floats; print them as integers rather
    than letting %g turn 2389040 into 2.38904e+06.
    """
    phv, n = cell
    if phv is None and n is None:
        return "-"

    def one(v):
        if v is None:
            return "-"
        if isinstance(v, float):
            return str(int(v)) if v.is_integer() else f"{v:g}"
        return str(v)

    return f"{one(phv)}/{one(n)}"
