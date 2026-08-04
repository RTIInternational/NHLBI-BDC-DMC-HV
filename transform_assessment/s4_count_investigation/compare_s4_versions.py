#!/usr/bin/env python
"""Compare Table S4 across dated exports of the Google Sheet.

Why this exists: the published S4 changed several times, and the changes do not
line up with changes to the transform specs.  Establishing *when* counts moved,
and whether label->count row alignment held across versions, is the entry point
to the count investigation.  ``compare_s4_to_published.py`` compares one
generated run against one published sheet; this compares published versions to
each other.

The sheets share a layout: labels in col 1, cohort headers in row 3, phv/n
subheaders in row 4, data from row 5, cohort column pairs from col 2.  The
2025-08-05 export has no TOTALS column; later ones carry TOTALS at cols 20/21.

Usage:
    ./.venv/bin/python transform_assessment/s4_count_investigation/compare_s4_versions.py
    ... --dir path/to/xslx --row "Alcohol Consumption"
    ... --labels          # row-alignment check only
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
DEFAULT_DIR = HERE / "xslx"
SHEET = "Table S4"
DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")


def version_key(path: Path) -> str:
    """Sort/label key: the date embedded in the filename, else the stem."""
    m = DATE_RE.search(path.name)
    return m.group(1) if m else path.stem


def load(path: Path) -> tuple[list[str], dict[str, dict[str, tuple]]]:
    """Return (labels in sheet order, {label: {cohort: (phv, n)}}).

    Labels are returned as a list as well as a dict so callers can check row
    *order* and duplication, not just membership -- a duplicated or shifted row
    is a live hypothesis for why published counts disagree with generated ones.
    """
    ws = openpyxl.load_workbook(path, data_only=True)[SHEET]
    cohorts = {
        ws.cell(row=3, column=c).value: c
        for c in range(2, ws.max_column + 1)
        if ws.cell(row=3, column=c).value
    }
    labels: list[str] = []
    rows: dict[str, dict[str, tuple]] = {}
    for r in range(5, ws.max_row + 1):
        raw = ws.cell(row=r, column=1).value
        if raw is None or not str(raw).strip():
            continue
        label = str(raw).strip()
        labels.append(label)
        rows.setdefault(
            label,
            {
                co: (ws.cell(row=r, column=c0).value, ws.cell(row=r, column=c0 + 1).value)
                for co, c0 in cohorts.items()
            },
        )
    return labels, rows


def fmt(cell: tuple) -> str:
    phv, n = cell
    return "-" if phv in (None, "") else f"{phv}/{n}"


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--dir", type=Path, default=DEFAULT_DIR, help="directory of .xlsx exports")
    ap.add_argument("--row", action="append", default=[], help="trace one label across versions (repeatable)")
    ap.add_argument("--labels", action="store_true", help="row-alignment check only")
    args = ap.parse_args()

    paths = sorted(args.dir.glob("*.xlsx"), key=version_key)
    if not paths:
        ap.error(f"no .xlsx files in {args.dir}")
    data = {version_key(p): load(p) for p in paths}
    versions = list(data)

    print("versions:", ", ".join(versions), end="\n\n")

    # Row alignment: duplicated labels within a version, and order changes
    # between versions.  A duplicate row is the signature of a bad paste --
    # counts were pasted positionally at cell B5 before xlsx generation existed,
    # so nothing joins a count to its label.
    print("== row alignment ==")
    for v in versions:
        labels, _ = data[v]
        dupes = {x for x in labels if labels.count(x) > 1}
        print(f"  {v}: {len(labels)} rows, {len(set(labels))} distinct" + (f"  DUPLICATED: {sorted(dupes)}" if dupes else ""))
    for a, b in zip(versions, versions[1:]):
        la, lb = data[a][0], data[b][0]
        shared = [x for x in la if x in set(lb)]
        misordered = [x for i, x in enumerate(shared) if [y for y in lb if y in set(la)][i] != x]
        note = f"{len(misordered)} shared labels in a different relative order" if misordered else "shared labels in the same order"
        print(f"  {a} -> {b}: {note}")
    print()

    if args.labels:
        return

    # Cell-level drift between consecutive versions.
    print("== cell changes between consecutive versions ==")
    for a, b in zip(versions, versions[1:]):
        _, A = data[a]
        _, B = data[b]
        common = set(A) & set(B)
        cohorts = (
            set().union(*[set(v) for v in A.values()]) & set().union(*[set(v) for v in B.values()])
        ) - {"TOTALS"}
        diffs = [(lab, co) for lab in common for co in cohorts if A[lab].get(co) != B[lab].get(co)]
        print(
            f"  {a} -> {b}: labels {len(A)}->{len(B)} (common {len(common)}), "
            f"cells differing {len(diffs)}/{len(common) * len(cohorts)}"
        )
    print()

    for target in args.row:
        print(f"== {target} ==")
        for v in versions:
            _, rows = data[v]
            row = rows.get(target)
            if row is None:
                near = [k for k in rows if target.lower()[:12] in k.lower()]
                print(f"  {v}: absent" + (f" (near: {near[:3]})" if near else ""))
                continue
            cells = "  ".join(f"{co}:{fmt(c)}" for co, c in row.items() if c[0] not in (None, ""))
            print(f"  {v}: {cells}")
        print()


if __name__ == "__main__":
    main()
