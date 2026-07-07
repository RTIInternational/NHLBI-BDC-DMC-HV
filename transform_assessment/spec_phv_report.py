#!/usr/bin/env python3
"""Spec-sourced Table S4: per-variable/cohort PHV counts and source N.

Replaces the spreadsheet-driven ``preharmonized_qaqc_report.py``. The
authoritative PHV->harmonized-variable mapping is the LinkML transform
specs in ``priority_variables_transform/<cohort>-ingest/<variable>.yaml``,
which have not drifted the way the Google Sheets have (e.g. SomaScan
aptamers mis-mapped to clinical concepts survive only in the sheets and in
``_archive/`` specs, not the live specs).

Two inputs, both spec/enclave-derived — no sheets:

1. **PHV list / count / harmonized variable** come from the specs. Each
   ``<variable>.yaml`` is loaded via linkml-map's normalizing loader
   (``load_specification``), which flattens the local ``observations``
   nesting into walkable ``class_derivations`` (see linkml/linkml-map issue
   #112). Every MeasurementObservation — flat, or nested inside a
   MeasurementObservationSet — contributes its own ``observation_type`` and
   the value-source PHVs under its value slots (``value_quantity`` /
   ``value_decimal`` / ``value_integer`` / ``value_concept`` /
   ``value_enum``). PHVs that only appear in ``associated_participant`` /
   ``associated_visit`` / ``age_at_observation`` are provenance, not value
   sources, and are excluded from the count.

   A spec file therefore yields **one row per distinct observation_type**:
   ``blood_pressure.yaml`` -> Systolic + Diastolic, ``spirometry.yaml`` ->
   FEV1 / FVC / FEV1-FVC / ..., each with its own PHVs and N. Concepts that
   resolve to the same label (whether across files, or a variable coded OBA
   in some cohorts and OMOP in others — HDL, LDL, triglycerides) merge back
   into a single row rather than splitting.

2. **N** comes from a source-extraction JSON produced by
   ``extract_source_summaries.py`` on SB, joined PHV -> column -> n_valid via
   the dbGaP cache's PHV name map. Run the source extract first (the
   ``run_extracts.sh``/``--yaml-dir`` path); this script consumes its JSON.

The label for each concept is resolved from ``harmonized_vars.tsv`` by its
``observation_type`` concept code (OMOP/OBA), falling back to the spec
filename stem treated as a ``var_name``, then the stem itself.

Usage::

    uv run python spec_phv_report.py \\
        --specs-root ../priority_variables_transform \\
        --source-json /path/to/<cohort>_source_*.json \\
        --cache-dir ../hv_dataqc/local_output/dbgap-cache/<cohort> \\
        --cohort FHS

Multiple ``--source-json``/``--cache-dir``/``--cohort`` triples may be
passed (repeat the flags) to build the full multi-cohort table in one run.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any

import yaml

from hv_dataqc.hv_dataqc_common import load_phv_name_map, XLSX_FMT_COUNT
from hv_dataqc.extract_harmonized.label_map import load_label_map

_PHV_RE = re.compile(r"phv\d+", re.IGNORECASE)

# Value slots whose source PHV(s) count as the harmonized variable's source
# variables.  PHVs appearing only outside these slots (participant id, visit,
# age) are provenance, not source measurements, and must not inflate the count.
_VALUE_SLOTS = ("value_decimal", "value_integer", "value_coded", "value_concept", "value_enum")

# Ingest directory name -> cohort label, mirroring run_extracts/run_s5.
_INGEST_DIR_RE = re.compile(r"^(?P<cohort>.+)-ingest$")


def _regroup_by_entity(path: Path) -> dict:
    """Regroup one spec file's blocks into the per-entity ``class_derivations``
    list form that ``linkml_map.load_specification`` expects.

    A per-variable spec file is a YAML list of ``{class_derivations: {Entity:
    ...}}`` blocks.  ``load_specification`` requires the composed dict form
    (``{class_derivations: [{Entity: ...}, ...]}``).  This is the same grouping
    dm-bip's ``compose_specs`` does, inlined here for one file so S4 needs
    neither the dm-bip package (whose pin conflicts with our linkml-map @main)
    nor a cohort-wide compose step.  Restore ``dm_bip.map_data.compose_specs``
    once dm-bip's linkml-map dependency catches up to the version carrying the
    nested-derivation fix (linkml/linkml-map d5abfd0).
    """
    blocks = yaml.safe_load(path.read_text()) or []
    if isinstance(blocks, dict):
        blocks = [blocks]
    entity_blocks: dict[str, list[dict]] = defaultdict(list)
    for block in blocks:
        cds = (block or {}).get("class_derivations")
        if not isinstance(cds, dict):
            continue
        for entity in cds:
            entity_blocks[entity].append(cds)
    return {"class_derivations": [{e: d[e]} for e, lst in entity_blocks.items() for d in lst]}


def _load_spec(path: Path):
    """Load a per-variable spec via linkml-map's normalizing loader.

    The loader normalizes the local ``observations``/``object_derivations``
    nesting into list-based ``class_derivations`` (linkml/linkml-map, see the
    deprecation of ``object_derivations``), so nested MeasurementObservations
    inside a MeasurementObservationSet become walkable rather than being
    dropped.
    """
    from linkml_map.utils.loaders import load_specification

    tmp = Path(tempfile.mktemp(suffix=".yaml"))
    tmp.write_text(yaml.dump(_regroup_by_entity(path), sort_keys=False))
    try:
        return load_specification(tmp)
    finally:
        tmp.unlink()


def _as_list(derivations) -> list:
    """ClassDerivation collections normalize to either a list or a name->cd
    dict depending on nesting depth; iterate values either way."""
    if not derivations:
        return []
    return derivations if isinstance(derivations, list) else list(derivations.values())


def _value_source_phvs(mo_sd: dict) -> dict[str, str | None]:
    """Map value-source PHV -> its PHT for one MeasurementObservation.

    *mo_sd* is a MeasurementObservation's ``slot_derivations`` dict from the
    typed spec.  Collects PHVs only from the value slots (and the
    ``value_quantity`` Quantity that wraps them).  The PHT is the
    ``populated_from`` on the nested Quantity (or the MO itself for flat slots)
    — needed to disambiguate columns whose name recurs across PHTs (e.g. "AST").
    """
    phv_pht: dict[str, str | None] = {}
    outer_pht = getattr(mo_sd.get("_mo"), "populated_from", None)

    vq = mo_sd.get("value_quantity")
    if vq is not None:
        for q in _as_list(getattr(vq, "class_derivations", None)):
            q_pht = getattr(q, "populated_from", None) or outer_pht
            inner = getattr(q, "slot_derivations", None) or {}
            for slot in _VALUE_SLOTS:
                sd = inner.get(slot)
                if sd is not None:
                    for phv in _slot_phvs(sd):
                        phv_pht.setdefault(phv, q_pht)

    # flat (non-nested) value slots directly on the MO
    for slot in _VALUE_SLOTS:
        sd = mo_sd.get(slot)
        if sd is not None:
            for phv in _slot_phvs(sd):
                phv_pht.setdefault(phv, outer_pht)

    return phv_pht


def _slot_phvs(slot_derivation) -> set[str]:
    """PHVs on one SlotDerivation: its ``populated_from`` plus any in ``expr``."""
    phvs: set[str] = set()
    pf = getattr(slot_derivation, "populated_from", None)
    if isinstance(pf, str):
        phvs.update(m.group(0).lower() for m in _PHV_RE.finditer(pf))
    expr = getattr(slot_derivation, "expr", None)
    if isinstance(expr, str):
        phvs.update(m.group(0).lower() for m in _PHV_RE.finditer(expr))
    return phvs


def _measurement_observations(spec) -> list:
    """Every MeasurementObservation slot_derivations dict in the spec — both
    flat top-level MOs and those nested under a MeasurementObservationSet's
    ``observations`` slot.  Each dict is tagged with its owning ClassDerivation
    under the ``_mo`` key so the PHT is recoverable."""
    out: list[dict] = []

    def collect(cd) -> None:
        sd = getattr(cd, "slot_derivations", None) or {}
        # A MeasurementObservationSet nests its MOs under `observations`.
        obs = sd.get("observations")
        nested = _as_list(getattr(obs, "class_derivations", None)) if obs is not None else []
        if nested:
            for mo in nested:
                collect(mo)
        else:
            tagged = dict(sd)
            tagged["_mo"] = cd
            out.append(tagged)

    for cd in _as_list(getattr(spec, "class_derivations", None)):
        collect(cd)
    return out


def parse_spec_file(path: Path) -> dict[str, Any]:
    """Return {variable, concepts} for one spec file.

    A spec file may define several distinct harmonized concepts (e.g.
    ``blood_pressure.yaml`` -> systolic + diastolic), each a MeasurementObservation
    with its own ``observation_type`` and value-source PHVs.  ``concepts`` is a
    list of ``{observation_type, phv_pht}``, one per distinct observation_type,
    unioning value-source PHVs across every MO (and PHT) that shares it.
    """
    spec = _load_spec(path)

    by_type: dict[str | None, dict[str, str | None]] = defaultdict(dict)
    for mo_sd in _measurement_observations(spec):
        ot = mo_sd.get("observation_type")
        otval = getattr(ot, "value", None) if ot is not None else None
        for phv, pht in _value_source_phvs(mo_sd).items():
            # First PHT wins; a PHV should map to one value column/PHT.
            by_type[otval].setdefault(phv, pht)

    concepts = [
        {"observation_type": otval, "phv_pht": phv_pht}
        for otval, phv_pht in by_type.items()
        if phv_pht
    ]
    return {"variable": path.stem, "concepts": concepts}


def load_var_labels(harmonized_vars_tsv: Path) -> dict[str, str]:
    """Map spec short name (var_name) -> publication label (var_label)."""
    labels: dict[str, str] = {}
    with harmonized_vars_tsv.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh, delimiter="\t")
        for row in reader:
            name = (row.get("var_name") or "").strip()
            label = (row.get("var_label") or "").strip()
            if name:
                labels[name] = label or name
    return labels


def _col_n_valid(
    source_doc: dict, phv: str, pht: str | None, phv_name_map: dict[str, str]
) -> int | None:
    """Resolve a PHV's n_valid from a source-extract JSON.

    PHV -> column name (via dbGaP cache map) -> n_valid in
    ``variables_by_pht[pht][column]``.  The PHT comes from the spec, which
    disambiguates column names that recur across PHTs (e.g. "AST" appears in
    several FHS PHTs).  JSON keys are lowercased column names; the cache map
    yields original case.  Falls back to scanning all PHTs only when the spec
    gave no PHT (should not happen for value slots).
    """
    col = (phv_name_map.get(phv) or phv).lower()
    vp = source_doc.get("variables_by_pht", {})

    if pht and pht in vp:
        stat = vp[pht].get(col)
        if isinstance(stat, dict) and stat.get("n_valid") is not None:
            return int(stat["n_valid"])
        return None

    # No PHT scope — sum unique matches as a last resort.
    total, found = 0, False
    for pht_cols in vp.values():
        stat = pht_cols.get(col)
        if isinstance(stat, dict) and stat.get("n_valid") is not None:
            total += int(stat["n_valid"])
            found = True
    return total if found else None


def _resolve_label(
    observation_type: str | None,
    stem: str,
    var_labels: dict[str, str],
    obs_type_labels: dict[str, str],
) -> str:
    """Resolve one concept's publication label.

    Order: by ``observation_type`` concept code (OMOP/OBA, the same join S5
    uses — robust to filename-vs-var_name drift), then by filename stem treated
    as a var_name, then the stem itself. This is what lets e.g. basophil_ct.yaml
    (obs_type OBA:VT0002607) resolve to "basophils count" instead of falling
    back to the raw stem.

    Resolving by concept code also collapses cross-vocabulary synonyms: a
    variable coded OBA in some cohorts and OMOP in others (HDL, LDL,
    triglycerides) lands on one label per code, so same-label concepts merge
    into a single row rather than splitting.  In every dual-coded case in the
    current specs it is the OBA code that resolves, giving an effective
    prefer-OBA / fall-back-OMOP behavior for free.
    """
    if observation_type and observation_type in obs_type_labels:
        return obs_type_labels[observation_type]
    return var_labels.get(stem, stem)


def build_cohort_rows(
    specs_dir: Path,
    source_json: Path,
    cache_dir: Path,
    var_labels: dict[str, str],
    obs_type_labels: dict[str, str] | None = None,
) -> dict[str, dict[str, Any]]:
    """Per-label {phv_count, total_n, phvs, label} for one cohort's specs.

    One spec file may contribute several rows (e.g. blood_pressure.yaml ->
    Systolic + Diastolic), one per distinct ``observation_type``.  Concepts
    resolving to the same label — whether across files or across vocabularies
    within a file — merge into one row (PHVs unioned, N summed once per PHV).
    Rows are keyed by resolved label; ``_build_table`` groups on label anyway.
    """
    source_doc = json.loads(source_json.read_text())
    phv_name_map = load_phv_name_map(cache_dir)
    obs_type_labels = obs_type_labels or {}

    # label -> merged {phv -> pht, observation_types set}
    merged: dict[str, dict[str, Any]] = {}
    for spec_path in sorted(specs_dir.glob("*.yaml")):
        parsed = parse_spec_file(spec_path)
        for concept in parsed["concepts"]:
            label = _resolve_label(
                concept["observation_type"], parsed["variable"], var_labels, obs_type_labels
            )
            slot = merged.setdefault(label, {"phv_pht": {}, "observation_types": set()})
            for phv, pht in concept["phv_pht"].items():
                slot["phv_pht"].setdefault(phv, pht)
            if concept["observation_type"]:
                slot["observation_types"].add(concept["observation_type"])

    rows: dict[str, dict[str, Any]] = {}
    for label, slot in merged.items():
        phv_pht = slot["phv_pht"]
        phvs = sorted(phv_pht)
        total_n = 0
        n_seen = False
        for phv in phvs:
            n = _col_n_valid(source_doc, phv, phv_pht[phv], phv_name_map)
            if n is not None:
                total_n += n
                n_seen = True
        rows[label] = {
            "label": label,
            "phv_count": len(phvs),
            "phvs": phvs,
            "phv_pht": phv_pht,
            "total_n": total_n if n_seen else None,
            "observation_type": sorted(slot["observation_types"]) or None,
        }
    return rows


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--specs-root", required=True, type=Path,
                   help="priority_variables_transform/ root containing <cohort>-ingest dirs.")
    p.add_argument("--harmonized-vars", type=Path,
                   default=Path(__file__).resolve().parents[1]
                   / "hv_dataqc/extract_harmonized/config/harmonized_vars.tsv",
                   help="harmonized_vars.tsv for var_name -> var_label labels.")
    p.add_argument("--cohort", action="append", required=True,
                   help="Cohort label; repeat per cohort (matches <cohort>-ingest dir).")
    p.add_argument("--source-json", action="append", required=True, type=Path,
                   help="Source-extract JSON for the matching --cohort; repeat in order.")
    p.add_argument("--cache-dir", action="append", required=True, type=Path,
                   help="dbGaP cache dir for the matching --cohort; repeat in order.")
    p.add_argument("--output", type=Path, default=Path(__file__).parent / "spec_phv_report.csv",
                   help="Output CSV path. A formatted .xlsx is also written alongside it "
                        "(same stem, .xlsx) unless --no-xlsx is given.")
    p.add_argument("--no-xlsx", action="store_true",
                   help="Do not also write the formatted .xlsx (CSV only).")
    p.add_argument("--layout", type=Path, default=None,
                   help=f"S4 layout config (canonical cohort columns + variable rows). "
                        f"Default: {DEFAULT_LAYOUT_PATH}")
    p.add_argument("--debug-variable", metavar="VAR",
                   help="Print the resolved PHVs + per-PHV n for this variable (short name).")
    args = p.parse_args(argv)

    if not (len(args.cohort) == len(args.source_json) == len(args.cache_dir)):
        p.error("--cohort, --source-json, --cache-dir must be repeated the same number of times")

    var_labels = load_var_labels(args.harmonized_vars)
    # observation_type concept code -> var_label, the robust join (S5 uses the
    # same map) that survives filename-stem vs var_name drift.
    obs_type_labels = load_label_map(args.harmonized_vars)

    # cohort -> {variable -> row}
    per_cohort: dict[str, dict[str, dict]] = {}
    for cohort, src, cache in zip(args.cohort, args.source_json, args.cache_dir):
        ingest_dir = args.specs_root / f"{cohort}-ingest"
        if not ingest_dir.is_dir():
            print(f"WARNING: no ingest dir for cohort {cohort}: {ingest_dir}", file=sys.stderr)
            continue
        per_cohort[cohort] = build_cohort_rows(ingest_dir, src, cache, var_labels, obs_type_labels)
        print(f"{cohort}: {len(per_cohort[cohort])} variables from {ingest_dir.name}")

    if args.debug_variable:
        _debug(args, var_labels, per_cohort)
        return 0

    layout = load_layout(args.layout)
    _write_csv(args.output, per_cohort, var_labels, layout)
    print(f"\nWrote {args.output}")
    if not args.no_xlsx:
        xlsx_path = args.output.with_suffix(".xlsx")
        _write_xlsx(xlsx_path, per_cohort, var_labels, layout)
        print(f"Wrote {xlsx_path}")
    return 0


def _debug(args, var_labels, per_cohort) -> None:
    var = args.debug_variable
    print(f"\n=== DEBUG variable={var!r} ({var_labels.get(var, var)}) ===")

    def find_row(rows: dict) -> dict | None:
        # Rows are keyed by resolved label; accept an exact key, the label a
        # var_name resolves to, or a case-insensitive label match.
        if var in rows:
            return rows[var]
        target = var_labels.get(var, var).strip().lower()
        for label, row in rows.items():
            if label.strip().lower() in (var.strip().lower(), target):
                return row
        return None

    for cohort, src, cache in zip(args.cohort, args.source_json, args.cache_dir):
        rows = per_cohort.get(cohort, {})
        row = find_row(rows)
        if not row:
            print(f"  {cohort}: (not present)")
            continue
        source_doc = json.loads(Path(src).read_text())
        phv_name_map = load_phv_name_map(Path(cache))
        print(f"  {cohort}: phv_count={row['phv_count']} total_n={row['total_n']} "
              f"obs_type={row['observation_type']}")
        for phv in row["phvs"]:
            col = phv_name_map.get(phv, phv)
            pht = row["phv_pht"].get(phv)
            n = _col_n_valid(source_doc, phv, pht, phv_name_map)
            print(f"      {phv}  pht={pht}  col={col!r}  n_valid={n}")


DEFAULT_LAYOUT_PATH = Path(__file__).parent / "config" / "s4_layout.yaml"


def load_layout(path: Path | None) -> dict:
    """Load the S4 layout config (canonical cohort columns + variable rows).

    Returns a dict with:
      - cohorts: ordered display cohort names (column order)
      - cohort_keys: {display_name -> internal cohort key} for names that
        differ between the template and the ingest-dir cohort (e.g.
        "HCHS/SOL" -> "HCHS"). Names not listed map to themselves.
      - variables: ordered template row labels (may be empty)
    Returns empty/defaults if the file is absent.
    """
    effective = path or DEFAULT_LAYOUT_PATH
    if not effective.exists():
        return {"cohorts": [], "cohort_keys": {}, "variables": []}
    cfg = yaml.safe_load(effective.read_text()) or {}
    return {
        "cohorts": list(cfg.get("cohorts") or []),
        "cohort_keys": dict(cfg.get("cohort_keys") or {}),
        "variables": list(cfg.get("variables") or []),
        "aliases": dict(cfg.get("aliases") or {}),
        "unmatched_note": cfg.get("unmatched_note") or "",
    }


def _build_label_resolver(template_labels: list[str], aliases: dict) -> dict[str, str]:
    """Map a normalized spec label -> the template label it belongs in.

    Matching is case-insensitive (keys are lowercased/stripped). Each template
    label maps to itself; each alias maps to its template label. Aliases win
    only for genuine spelling differences — case is already handled by the
    normalization, so plain capitalization variants need no alias entry.
    """
    def norm(s: str) -> str:
        return str(s).strip().lower()

    resolver: dict[str, str] = {}
    for tpl in template_labels:
        resolver[norm(tpl)] = tpl
    for tpl, alts in aliases.items():
        for alt in alts or []:
            resolver[norm(alt)] = tpl
    return resolver


def _build_table(per_cohort, var_labels, layout: dict | None = None) -> tuple[list[str], list[list], list[str]]:
    """Build the S4 table as (headers, rows) — shared by the CSV and xlsx writers.

    Rows are value lists aligned to headers. When a canonical variable list is
    configured, rows follow it (blank where no data), then unmatched spec
    variables are appended after a blank separator + note row.
    """
    layout = layout or {}

    if layout.get("cohorts"):
        display_cohorts = layout["cohorts"]
        key_for = lambda disp: layout.get("cohort_keys", {}).get(disp, disp)
    else:
        display_cohorts = sorted(per_cohort)
        key_for = lambda disp: disp

    by_label: dict[str, dict[str, dict]] = {}
    for ckey, rows in per_cohort.items():
        for short, row in rows.items():
            by_label.setdefault(row.get("label", short), {})[ckey] = row

    headers = ["variable"]
    for disp in display_cohorts:
        headers += [f"{disp}_phv", f"{disp}_n"]
    headers += ["TOTALS_phv", "TOTALS_n"]  # row totals across cohorts

    def cells_for(label, label_rows: dict) -> list:
        out = [label]
        tot_phv = 0
        tot_n = 0
        any_data = False
        for disp in display_cohorts:
            row = label_rows.get(key_for(disp))
            phv = row["phv_count"] if row else ""
            n = row["total_n"] if row and row["total_n"] is not None else ""
            out.append(phv)
            out.append(n)
            if isinstance(phv, (int, float)):
                tot_phv += phv
                any_data = True
            if isinstance(n, (int, float)):
                tot_n += n
                any_data = True
        out.append(tot_phv if any_data else "")
        out.append(tot_n if any_data else "")
        return out

    template_labels = layout.get("variables") or []
    out_rows: list[list] = []

    if not template_labels:
        for label in sorted(by_label):
            out_rows.append(cells_for(label, by_label[label]))
        return headers, out_rows, []

    resolver = _build_label_resolver(template_labels, layout.get("aliases") or {})
    matched: dict[str, dict[str, dict]] = {}
    unmatched: list[str] = []
    for spec_label, cohort_rows in by_label.items():
        tpl = resolver.get(spec_label.strip().lower())
        if tpl is None:
            unmatched.append(spec_label)
            continue
        matched.setdefault(tpl, {}).update(cohort_rows)

    for label in template_labels:
        out_rows.append(cells_for(label, matched.get(label, {})))

    if unmatched:
        blank = [""] * len(headers)
        out_rows.append(blank)
        note = layout.get("unmatched_note") or (
            "Variables below have spec data but no matching Table S4 row."
        )
        out_rows.append([note] + [""] * (len(headers) - 1))
        for spec_label in sorted(unmatched):
            out_rows.append(cells_for(spec_label, by_label[spec_label]))

    return headers, out_rows, sorted(unmatched)


def _write_csv(output: Path, per_cohort, var_labels, layout: dict | None = None) -> None:
    headers, rows, unmatched = _build_table(per_cohort, var_labels, layout)
    with output.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        writer.writerows(rows)
    if unmatched:
        print(
            f"NOTE: {len(unmatched)} spec variable(s) had no template row; "
            f"appended at the bottom: {', '.join(unmatched)}",
            file=sys.stderr,
        )


_S4_TITLE = "Table S4: Raw variables and sample sizes by priority variable before harmonization"
_S4_CAPTION = "Number of relevant raw variables (phv) and data points (excluding missing values)"


def _write_xlsx(output: Path, per_cohort, var_labels, layout: dict | None = None) -> None:
    """Write the S4 xlsx replicating the template's 4-row merged header.

    Header rows:
      1. title (merged across all columns)
      2. "Priority Variable" (A) | caption (merged across the cohort columns)
      3. cohort names, each merged across its phv/n pair (+ TOTALS)
      4. "phv" / "n" under each group
      5+ data; counts get the #,##0 format.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment

    headers, rows, _ = _build_table(per_cohort, var_labels, layout)
    # Column groups: variable (1 col) then pairs. The header labels are derived
    # from the "<group>_phv"/"<group>_n" header names built in _build_table.
    groups: list[str] = []
    for h in headers[1:]:
        if h.endswith("_phv"):
            groups.append(h[:-4])
    ncols = len(headers)
    last_col = get_column_letter(ncols)

    wb = Workbook()
    ws = wb.active
    ws.title = "Table S4"
    center = Alignment(horizontal="center")

    # Row 1: title across everything.
    ws["A1"] = _S4_TITLE
    ws.merge_cells(f"A1:{last_col}1")
    # Row 2: A = "Priority Variable", caption spans the data columns.
    ws["A2"] = "Priority Variable"
    ws["B2"] = _S4_CAPTION
    ws.merge_cells(f"B2:{last_col}2")
    ws["B2"].alignment = center
    # Row 3: group names, each merged over its phv/n pair.
    # Row 4: phv / n labels.
    col = 2
    for g in groups:
        gl, gr = get_column_letter(col), get_column_letter(col + 1)
        ws[f"{gl}3"] = g
        ws.merge_cells(f"{gl}3:{gr}3")
        ws[f"{gl}3"].alignment = center
        ws[f"{gl}4"] = "phv"
        ws[f"{gr}4"] = "n"
        col += 2

    # Data rows start at row 5; numeric count columns get the count format.
    from hv_dataqc.hv_dataqc_common import coerce_number
    for r in rows:
        ws.append(r)
        excel_row = ws.max_row
        for ci in range(2, ncols + 1):  # skip the variable label column
            cell = ws.cell(row=excel_row, column=ci)
            val = coerce_number(cell.value)
            cell.value = val
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = XLSX_FMT_COUNT

    # Column widths + freeze the 4-row header.
    ws.column_dimensions["A"].width = 40
    for ci in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.freeze_panes = "B5"

    output.parent.mkdir(parents=True, exist_ok=True)
    tmp = output.with_name(f"{output.name}.tmp")
    try:
        wb.save(tmp)
        tmp.replace(output)
    finally:
        if tmp.exists():
            tmp.unlink()


if __name__ == "__main__":
    raise SystemExit(main())
